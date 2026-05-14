from fastapi import FastAPI, UploadFile, File, Form
from fastapi.middleware.cors import CORSMiddleware
from pypdf import PdfReader
from openpyxl import load_workbook
from dotenv import load_dotenv
import google.generativeai as genai
import os
import shutil

load_dotenv()

GEMINI_API_KEY = os.getenv("GEMINI_API_KEY")

if GEMINI_API_KEY:
    genai.configure(api_key=GEMINI_API_KEY)

app = FastAPI(title="Kairos MA-thos API")

app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "http://localhost:5500",
        "http://127.0.0.1:5500"
    ],
    allow_credentials=False,
    allow_methods=["*"],
    allow_headers=["*"],
)

UPLOAD_FOLDER = "uploads"
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

documentos = []
ultimo_documento_texto = ""
ultimo_documento_nome = ""

def ler_pdf(caminho):
    texto = ""
    reader = PdfReader(caminho)

    for pagina in reader.pages:
        conteudo = pagina.extract_text()
        if conteudo:
            texto += conteudo + "\n"

    return texto


def ler_xlsx(caminho):
    texto = ""
    workbook = load_workbook(caminho, data_only=True)

    for sheet in workbook.sheetnames:
        aba = workbook[sheet]
        texto += f"\nAba: {sheet}\n"

        for linha in aba.iter_rows(values_only=True):
            valores = [str(celula) for celula in linha if celula is not None]
            if valores:
                texto += " | ".join(valores) + "\n"

    return texto


def ler_txt(caminho):
    with open(caminho, "r", encoding="utf-8", errors="ignore") as arquivo:
        return arquivo.read()


def chamar_ia(texto, tipo_analise):
    if not GEMINI_API_KEY:
        return "Erro: chave GEMINI_API_KEY não configurada no arquivo .env."

    prompt = f"""
Você é uma IA de leitura inteligente chamada Pathos, do projeto Kairos MA-thos.

Analise o documento abaixo e faça a seguinte tarefa:

Tarefa: {tipo_analise}

Documento:
{texto[:25000]}
"""

    model = genai.GenerativeModel("gemini-2.5-flash")
    resposta = model.generate_content(prompt)

    return resposta.text


@app.get("/")
def home():
    return {"mensagem": "API Kairos MA-thos funcionando"}


@app.post("/upload")
async def upload_arquivo(
    arquivo: UploadFile = File(...),
    tipo_analise: str = Form(...)
):
    try:
        caminho = os.path.join(UPLOAD_FOLDER, arquivo.filename)

        with open(caminho, "wb") as buffer:
            shutil.copyfileobj(arquivo.file, buffer)

        extensao = arquivo.filename.lower().split(".")[-1]

        if extensao == "pdf":
            texto = ler_pdf(caminho)
        elif extensao == "xlsx":
            texto = ler_xlsx(caminho)
        elif extensao == "txt":
            texto = ler_txt(caminho)
        else:
            return {"erro": "Formato não suportado. Use PDF, XLSX ou TXT."}

        if not texto.strip():
            return {"erro": "Não foi possível extrair texto desse arquivo."}

        global ultimo_documento_texto, ultimo_documento_nome

        ultimo_documento_texto = texto
        ultimo_documento_nome = arquivo.filename

        resposta_ia = chamar_ia(texto, tipo_analise)

        item = {
            "nome": arquivo.filename,
            "tipo_analise": tipo_analise,
            "resposta": resposta_ia
        }

        documentos.append(item)

        return item

    except Exception as erro:
        return {
            "erro": "Erro interno no backend.",
            "detalhes": str(erro)
        }
    caminho = os.path.join(UPLOAD_FOLDER, arquivo.filename)

    with open(caminho, "wb") as buffer:
        shutil.copyfileobj(arquivo.file, buffer)

    extensao = arquivo.filename.lower().split(".")[-1]

    if extensao == "pdf":
        texto = ler_pdf(caminho)
    elif extensao == "xlsx":
        texto = ler_xlsx(caminho)
    elif extensao == "txt":
        texto = ler_txt(caminho)
    else:
        return {"erro": "Formato não suportado. Use PDF, XLSX ou TXT."}

    resposta_ia = chamar_ia(texto, tipo_analise)

    item = {
        "nome": arquivo.filename,
        "tipo_analise": tipo_analise,
        "resposta": resposta_ia
    }

    documentos.append(item)

    return item


@app.get("/historico")
def listar_historico():
    return documentos

@app.post("/chat")
async def chat_documento(pergunta: str = Form(...)):
    try:
        if not ultimo_documento_texto.strip():
            return {
                "erro": "Nenhum documento foi carregado ainda. Envie um documento primeiro."
            }

        prompt = f"""
Você é uma IA especialista em leitura e interpretação de documentos.

Regras:
- Responda de forma clara e organizada
- Use linguagem simples
- Se possível, responda em tópicos
- NÃO invente informações
- Se não encontrar a resposta no texto, diga claramente que não encontrou

Documento:
{ultimo_documento_texto[:25000]}

Pergunta:
{pergunta}
"""

        model = genai.GenerativeModel("gemini-2.5-flash")
        resposta = model.generate_content(prompt)

        return {
            "arquivo": ultimo_documento_nome,
            "pergunta": pergunta,
            "resposta": resposta.text
        }

    except Exception as erro:
        return {
            "erro": "Erro ao conversar com o documento.",
            "detalhes": str(erro)
        }